import openpyxl
import os
import re
os.chdir('C:/Users/MyFirstBuild/AppData/Local/Programs/Python/Python37')
wb = openpyxl.load_workbook('challenge_1.xlsx')
sheet = wb.get_sheet_by_name('Form Responses 1')
newPattern =re.compile(r'\d{10}')
for i in range(2, 100):
	if(newPattern.search(str(sheet.cell(row=i, column=8).value)) != None):
		sheet.cell(row=i, column=9).value=int(newPattern.findall(str(sheet.cell(row=i, column=8).value))[0])
ws=wb.active
for i in range(2, 100):
	   if(sheet.cell(row=i, column=4).value == sheet.cell(row=i-1, column=4).value):
		   ws.delete_rows(i)
wb.save('challenge_1_complete.xlsx')
